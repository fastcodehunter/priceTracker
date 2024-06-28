import datetime
import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import json

def scrapy_data(lst_url):
    print('Cбор данных')
    dict_data1 = [{'Model': '', 'Shop': []}]
    pattern = r'gree/([^/]+)'
    for title,model in lst_url.items():
        for color in model:
            url = f'https://catalog.onliner.by/conditioners/gree/{color}/prices'
            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')
            atribut_a = soup.find_all('a', class_='offers-description-filter-control offers-description-filter-control_switcher js-facet-configurations-link')
            try:
             atribut_color = soup.find('div', class_='offers-description-filter__value').get_text().strip().upper()

            except AttributeError:
                atribut_color = ""
               
            all_power = [power['href'] for power in atribut_a]
            all_power.append(url)
            
            for power in all_power:
                response = requests.get(power)
                soup = BeautifulSoup(response.content, 'html.parser')
                title_model=soup.find('h1').get_text().strip().replace('Цены на сплит-систему ','')

                print(title_model)
                match = re.search(pattern, power).group(1)
                if match.endswith('dna1dw'):
                    match = match[:-1]
                json_url = f'https://catalog.onliner.by/sdapi/shop.api/products/{match}/positions?town=all&has_prime_delivery=1&town_id=17030'
                
                get_json_data = requests.get(json_url)
                json_html = get_json_data.json()
                
                model_data = {'Model': f"{title_model}", 'Shop': []}
                for shop_id, price in zip(json_html['shops'], json_html['positions']['primary']):
                    model_data['Shop'].append({
                        'Shop': json_html['shops'][shop_id]['title'],
                        'Price': f"{price['position_price']['amount']} {price['position_price']['currency']}"
                    })
                dict_data1.append(model_data)
    return dict_data1         

lst_url = { 
            'pular': ['gwh09agaxak6dna4'],
            'bora': ['gwh09aaaxak6dna2'],
            'arctic': ['gwh09qcxbk6dnc2f'],
            'r410':['gwh09agaxak3nna1'],
            'console':['geh09aak6dna1f'],
            'lyra':['gwh12acck6dna1fw','gwh12acck6dna1f','gwh12acck6dna1fh'],
            'airy':['gwh09avcxbk6dnaw','gwh09avcxbk6dnab','gwh09avcxbk6dnac'],


}



workbook = openpyxl.Workbook()
sheet = workbook.active

sheet['A1'] = 'Модель'
sheet['B1'] = 'Название магазина'
sheet['C1'] = 'Цена'

row = 2
for conditioner in scrapy_data(lst_url):
    sheet.cell(row=row, column=1, value=(conditioner['Model']))
    for shop in conditioner['Shop']:
        sheet.cell(row=row, column=2, value=shop['Shop'])
        sheet.cell(row=row, column=3, value=shop['Price'])
        row += 1
    row += 1

workbook.save('conditioner.xlsx')
print("Загрузка завершена!")



